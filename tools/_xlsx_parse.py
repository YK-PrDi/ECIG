# -*- coding: utf-8 -*-
"""读 xlsx → 抽出每个品类下的字段选项与卖点。

xlsx 约定（参考"人工转IT字段流程"系列）：
- 第 1 行（C1..Cn）= 列头，其中 C1 = 品类路径，C2..Cn = 字段中文名（如 '安装方式' '形态结构' '卖点' '场景构图' ...）
- 第 2..N 行 = 数据；每行可视作"一组卖点+对应字段值"。同一品类的所有行属于同一品类。
- 'C1 品类' 通常只在第 2 行写出，其余行留空表示沿用上一行品类。

输出字段名（中文）→ fields.js 里 EC_FIELDS 的 key 通过 promptKey 反查得到。
'场景构图'（C 列在 xlsx 7/8 里）→ 同一行的"卖点 prompt + composition"组合：
  卖点的 prompt = xlsx 同行 '卖点' 列原值 + 字体模板（如有）
  卖点的 composition = 同行 '场景构图' + '拍摄角度' + '产品摆放' + '场景布局' + '产品状态' + '产品质感' + '光影'
  这些"分镜级别"字段还会同时被合并到本品类的字段选项池里（作为下次可勾选的预设）。
"""
from __future__ import annotations
import re
from pathlib import Path
from typing import Iterable

import openpyxl


# xlsx 中文列头 → fields.js 里的 promptKey（即"字段中文名"）
# fields.js 的 promptKey 就是这些中文，所以直接同名映射即可。
HEADER_ALIASES = {
    '品类': '__category__',
    '图片类型': '图片类型',
    '简短提示词': '附加拍摄指令',
    '附加拍摄指令': '附加拍摄指令',
    '安装方式': '安装方式',
    '按安装方式': '安装方式',
    '形态结构': '形态',
    '按形态结构': '形态',
    '形态': '形态',
    '背景风格': '背景风格',
    '材质/工艺': '材质/工艺',
    '材质': '材质/工艺',
    '参考图': '参考图',
    '卖点': '__selling__',          # 特殊：行级别拆成卖点对象
    '主图卖点': '__selling__',      # xlsx 10 用了"主图卖点【核心竞争力】（卖点差异化）"
    '主图卖点核心竞争力': '__selling__',
    '字体模板': '字体模板',
    '场景构图': '__composition__',  # 特殊：进入卖点 composition
    '拍摄角度': '拍摄角度',
    '产品摆放': '产品摆放',
    '场景布局': '场景布局',
    '产品状态': '产品状态',
    '画质': '画质',
    '产品质感': '产品质感',
    '光影': '光影',
    '场景': '场景',
    '场景细节': '场景细节',
    '滤芯个数': '滤芯个数',
    '滤芯白底图': '__skip__',  # 行级补充信息，不进入字段或卖点
}

# 全局字段（在 fields.js 标 global: true）：选项由 fields.js 提供默认；品类如有同名列也会注入到品类专属
# 这里仅列出"完全不允许品类自定义"的全局字段（如图片类型、画质这种通用枚举）
GLOBAL_FIELDS = {'图片类型', '参考图', '画质', '附加拍摄指令', '字体模板', '滤芯个数'}

# Excel 错误值，原样导入会污染下拉
EXCEL_ERRORS = {'#NAME?', '#REF!', '#VALUE!', '#DIV/0!', '#N/A', '#NULL!', '#NUM!'}

# 占位符 / 无意义值：xlsx 编辑期会写"自定义""—""无""无说明""TBD"等占位，不应进入下拉选项
PLACEHOLDER_VALUES = {'自定义', '—', '-', '无', '无说明', '待补', 'TBD', '/', 'N/A', 'na', 'null'}

# 卖点列里偶尔出现的"工作备注式文字"——不是真卖点，要过滤掉
_SELLING_NOTE_KEYWORDS = ('我建议', '我觉得', '建议在', '可以选择', '注：', '备注：', '注意：', 'TODO', 'todo', 'sku制作', 'SKU制作', '可以同时', '同时选择')


def _is_placeholder(s: str) -> bool:
    if not s:
        return True
    s = s.strip()
    return s in PLACEHOLDER_VALUES or _is_excel_error(s)


def _looks_like_note(s: str) -> bool:
    """识别明显是工作备注/批注而非真正卖点的文本（会出现在 xlsx 卖点列里的人工说明）"""
    if not s:
        return True
    s = s.strip()
    return any(s.startswith(k) or k in s[:6] for k in _SELLING_NOTE_KEYWORDS)


def _is_excel_error(s: str) -> bool:
    return s in EXCEL_ERRORS or (isinstance(s, str) and s.startswith('#') and s.endswith(('?', '!')) and len(s) <= 10)


def _canon_header(s: str) -> str:
    """剥掉表头里的备注，如 '背景风格【视觉机调】' → '背景风格'。"""
    if not s:
        return ''
    s = re.sub(r'【[^】]*】', '', str(s))
    s = re.sub(r'\([^)]*\)|（[^）]*）', '', s)
    return s.strip()


def _normalize_ws(s: str) -> str:
    """折叠连续空白、去首尾空白；保留中文标点。"""
    if not s:
        return ''
    return re.sub(r'\s+', ' ', str(s)).strip()


def _split_selling(text: str) -> tuple[str, str]:
    """xlsx '卖点' 单元格清洗：
    - 形式 1：'主标语（布局说明）' → label=主标语, layout_note=布局说明
    - 形式 2：长段落，label 多余 30 字时用启发式精简：
        a) 优先取首个 双引号/中文引号 内的内容（用户常以引号包裹标语）
        b) 否则取 第一个 ；; 。\\n 1. 项目编号 之前的内容
      被截掉的部分进入 layout_note
    - 形式 3：短文本（≤30 字）原样返回
    """
    text = _normalize_ws(text)
    # 形式 1a：'主标语（布局说明）'（说明在末尾，整段 ≤ 60 字才算典型）
    m = re.match(r'^(.*?)\s*[（(]([^）)]*)[）)]\s*$', text)
    if m and len(m.group(1).strip()) <= 30:
        return m.group(1).strip(), m.group(2).strip()

    # 短文本直接用——但若含明显"已被截断/装饰说明"特征（…/未配对引号/分号+副标题），仍走启发式重洗
    looks_dirty = (
        '…' in text or '...' in text
        or text.count('"') == 1 or text.count('"') % 2 == 1
        or '副标题' in text or '主标题' in text or '作主标题' in text
    )
    if len(text) <= 30 and not looks_dirty:
        return text, ''

    # 长文本启发式：优先用最具结构性的信号
    # 2a) 引号内的标语（中英文双引号，2-30 字内）—— 用户常用引号包裹标语
    qm = re.search(r'["“”]([^"“”]{2,30})["“”]', text)
    if qm:
        label = qm.group(1).strip()
        rest = (text[:qm.start()] + ' ' + text[qm.end():]).strip()
        return label, rest

    # 2b) '主标语（说明）；其它...'：括号开始位置即分割点（短主标语场景）
    m = re.match(r'^(.{2,30}?)\s*[（(]', text)
    if m:
        label = m.group(1).strip()
        return label, text[m.end()-1:].strip()

    # 2c) 第一个明显分隔符前的部分
    for delim in ['\n', '；', ';', '。']:
        idx = text.find(delim)
        if 0 < idx <= 30:
            return text[:idx].strip(), text[idx + 1:].strip()

    # 2d) 项目编号 "1." "1、" 之前
    nm = re.match(r'^(.{2,30}?)\s*[12３4]\s*[\.\、]', text)
    if nm:
        return nm.group(1).strip(), text[nm.end() - 2:].strip()

    # 兜底：硬截 28 字 + …（_writer 会再做 short_display）
    return text[:28].rstrip() + '…', text


def parse_xlsx(path: Path) -> dict:
    """解析单个 xlsx 文件，返回：
    {
      '<完整品类路径>': {
         'fields': { '<中文字段名>': set(values) },
         'sellings': [ {label, prompt, composition} ]
      }
    }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    # 表头映射 col_idx → 字段中文名
    raw_headers = [_canon_header(c.value) for c in ws[1]]
    headers = [HEADER_ALIASES.get(h, '') for h in raw_headers]

    # 找各特殊列下标
    def col_idx(field: str) -> int:
        for i, h in enumerate(headers):
            if h == field:
                return i
        return -1

    cat_col      = col_idx('__category__')
    selling_col  = col_idx('__selling__')
    comp_col     = col_idx('__composition__')
    font_col     = col_idx('字体模板')
    if cat_col < 0:
        raise RuntimeError(f"xlsx {path.name} 找不到 '品类' 列（C1）")

    out: dict = {}
    last_cat = ''
    # composition 由这些字段串联（按出现顺序）
    composition_keys = ['场景构图', '拍摄角度', '产品摆放', '场景布局', '产品状态', '产品质感', '光影']

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = [str(c).strip() if c is not None else '' for c in row]
        if not any(cells):
            continue
        cat = cells[cat_col] if cat_col < len(cells) else ''
        if cat:
            last_cat = cat
        if not last_cat:
            continue
        # 用 ">" 标准化分隔符：xlsx 里有的写 "A>B"，有的写 "A——B"（中文双破折号），
        # 还有 "A&gt;B"（HTML 实体）/ "A->B"。catalog.js 的 ecCatalogResolveUp 只用 ">"
        # 拆/拼，所以这里必须把所有别名都归一成 ">"，否则会生成跟前端查找路径不匹配的
        # EC_CATALOG key —— 数据写进去了但前端永远找不到。
        cat_path = re.sub(r'\s*>\s*|\s*&gt;\s*|\s*->\s*|\s*——\s*|\s*——\s*|\s*-\s*', '>', last_cat).strip()
        bucket = out.setdefault(cat_path, {'fields': {}, 'sellings': []})

        # 1) 把每列值塞到 fields 池
        for i, field in enumerate(headers):
            if not field or field.startswith('__'):
                continue
            if field in GLOBAL_FIELDS:
                continue  # 全局字段不进品类专属
            if i >= len(cells):
                continue
            v = cells[i]
            if not v or _is_placeholder(v):
                continue
            bucket['fields'].setdefault(field, []).append(v)

        # 2) 卖点：当前行有 '卖点' 值才算一个卖点
        selling = cells[selling_col] if 0 <= selling_col < len(cells) else ''
        if selling and not _is_placeholder(selling) and not _looks_like_note(selling):
            label, layout_note = _split_selling(selling)
            comp_parts = []
            # 主 composition
            if 0 <= comp_col < len(cells) and cells[comp_col] and not _is_placeholder(cells[comp_col]):
                comp_parts.append(_normalize_ws(cells[comp_col]))
            # 把同行其它 composition 类字段也并进 composition 文本
            for k in composition_keys[1:]:
                ci = col_idx(k)
                if 0 <= ci < len(cells) and cells[ci] and not _is_placeholder(cells[ci]):
                    comp_parts.append(f"{k}：{_normalize_ws(cells[ci])}")
            font = cells[font_col] if 0 <= font_col < len(cells) else ''
            if _is_placeholder(font):
                font = ''
            font = _normalize_ws(font)
            # prompt = 主标语 + （布局说明）+ ；文字模板：xxx
            prompt = label
            if layout_note:
                prompt += f'（{layout_note}）'
            if font:
                prompt += f'；文字模板：{font}'
            bucket['sellings'].append({
                'label': label,
                'prompt': prompt,
                'composition': '；'.join(comp_parts) if comp_parts else '',
            })

    return out


def parse_many(paths: Iterable[Path]) -> dict:
    """合并多个 xlsx 输出（同品类合并，字段选项 set 取并集，卖点 label 去重）。"""
    merged: dict = {}
    for p in paths:
        single = parse_xlsx(p)
        for cat, data in single.items():
            target = merged.setdefault(cat, {'fields': {}, 'sellings': []})
            for k, vs in data['fields'].items():
                pool = target['fields'].setdefault(k, [])
                for v in vs:
                    if v not in pool:
                        pool.append(v)
            seen_labels = {s['label'] for s in target['sellings']}
            for s in data['sellings']:
                if s['label'] not in seen_labels:
                    target['sellings'].append(s)
                    seen_labels.add(s['label'])
    return merged
